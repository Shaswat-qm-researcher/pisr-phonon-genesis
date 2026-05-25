#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Cv vs T Comparison: Phonopy data vs SISSO Debye model vs Fitted Debye model.

Inputs:
Chemical formula from PRIMCELL.vasp (primitive cell).
N_ATOMS is taken from thermal_properties.yaml (natom field) so the Debye model and...
... Dulong-Petit limit are always consistent with the actual supercell used in phonopy.

Outputs:
    Cv_comparison_<formula>.pdf   
    Cv_comparison_<formula>.tiff  
    Cv_comparison_<formula>_source_data.xlsx

Requirements:
    pip install numpy scipy pyyaml matplotlib phonopy openpyxl Pillow
"""

import os
import sys
import math
import yaml
import numpy as np
from scipy.integrate import quad
from scipy.optimize import curve_fit
from collections import OrderedDict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker

plt.rcParams.update({
    # Font
    'font.family'       : 'sans-serif',
    'font.sans-serif'   : ['Arial', 'DejaVu Sans'],
    'font.size'         : 7,            # base size
    'mathtext.fontset'  : 'dejavusans',

    # Axes
    'axes.linewidth'    : 0.75,         
    'axes.labelsize'    : 7,            # axis label font size (pt)
    'axes.titlesize'    : 7,

    # Ticks
    'xtick.major.width' : 0.75,
    'ytick.major.width' : 0.75,
    'xtick.minor.width' : 0.5,
    'ytick.minor.width' : 0.5,
    'xtick.major.size'  : 3.5,
    'ytick.major.size'  : 3.5,
    'xtick.minor.size'  : 2.0,
    'ytick.minor.size'  : 2.0,
    'xtick.direction'   : 'in',
    'ytick.direction'   : 'in',
    'xtick.labelsize'   : 6,
    'ytick.labelsize'   : 6,

    # Legend
    'legend.fontsize'   : 6,
    'legend.framealpha' : 0.55,

    # Lines
    'lines.linewidth'   : 1.0,

    # Output / font embedding
    'pdf.fonttype'      : 42,           # embed TrueType (required by Nature)
    'ps.fonttype'       : 42,
    'svg.fonttype'      : 'none',
})

# ── Nature column widths (inches) ─────────────────────────────────────────────
_MM_TO_IN = 1.0 / 25.4
NATURE_SINGLE_COL_W = 89  * _MM_TO_IN   # 3.504 in
NATURE_DOUBLE_COL_W = 183 * _MM_TO_IN   # 7.205 in
NATURE_MAX_HEIGHT   = 247 * _MM_TO_IN   # 9.724 in

R = 8.314462618   # J / (mol·K)

# ─────────────────────────────────────────────────────────────────────────────
# Debye model
# ─────────────────────────────────────────────────────────────────────────────
def _debye_integrand(x: float) -> float:
    if x < 1e-12:
        return 0.0
    emx = math.exp(-x)
    denom = (1.0 - emx) ** 2
    return x**4 * emx / denom if denom > 1e-300 else 0.0


def debye_Cv_single(T: float, theta_D: float, n_atoms: int) -> float:
    if T <= 0:
        return 0.0
    xD  = theta_D / T
    val, _ = quad(_debye_integrand, 0.0, xD, limit=200)
    return 9.0 * n_atoms * R * (T / theta_D)**3 * val


def debye_Cv_array(T_arr: np.ndarray, theta_D: float,
                   n_atoms: int) -> np.ndarray:
    return np.array([debye_Cv_single(t, theta_D, n_atoms) for t in T_arr])


# ─────────────────────────────────────────────────────────────────────────────
# PRIMCELL.vasp / POSCAR parser  →  formula strings
# ─────────────────────────────────────────────────────────────────────────────
def parse_vasp_species(filepath: str) -> OrderedDict:
    """
    Parse species and counts from any VASP5-format structure file.
    Returns OrderedDict {element: count}.
    """
    with open(filepath) as fh:
        lines = [l.rstrip('\n') for l in fh]

    tok5 = lines[5].split()
    tok6 = lines[6].split() if len(lines) > 6 else []

    if not tok5 or not tok5[0].isalpha():
        raise ValueError(
            f"'{filepath}' looks like a VASP4 file (no species names on line 6). "
            "Please use a VASP5-format file."
        )

    species = tok5
    counts  = [int(x) for x in tok6]
    if len(species) != len(counts):
        raise ValueError(
            f"Species/count mismatch in '{filepath}': {species} vs {counts}"
        )
    return OrderedDict(zip(species, counts))


def build_formula_strings(species: OrderedDict) -> tuple[str, str]:
    """Return (plain_formula, latex_formula), e.g. ('B13CN', 'B$_{13}$CN')."""
    plain = "".join(
        f"{s}{c}" if c > 1 else s for s, c in species.items()
    )
    latex = "".join(
        f"{s}$_{{{c}}}$" if c > 1 else s for s, c in species.items()
    )
    return plain, latex


# ─────────────────────────────────────────────────────────────────────────────
# thermal_properties.yaml reader
# ─────────────────────────────────────────────────────────────────────────────
def load_cv_from_yaml(yaml_file: str) -> tuple[np.ndarray, np.ndarray, int]:
    """Return (T_array, Cv_array, natom) from phonopy thermal_properties.yaml."""
    with open(yaml_file) as fh:
        data = yaml.safe_load(fh)
    temps, cvs = [], []
    for tp in data["thermal_properties"]:
        temps.append(float(tp["temperature"]))
        cvs.append(float(tp["heat_capacity"]))
    natom = int(data.get("natom", 0))
    return np.array(temps), np.array(cvs), natom


# ─────────────────────────────────────────────────────────────────────────────
# Directory prompt
# ─────────────────────────────────────────────────────────────────────────────
def prompt_dir(label: str, default: str = None) -> str:
    prompt = f"{label} [{default}]: " if default else f"{label}: "
    while True:
        val = input(prompt).strip() or default or ''
        val = os.path.expanduser(val)
        if os.path.isdir(val):
            return val
        print(f"  ✗ Not found: '{val}'. Try again.")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    print("=" * 62)
    print("   Cv vs T  ·  Phonopy / SISSO / Fitted Debye  ·  Nature Quality")
    print("=" * 62)

    # User settings
    THETA_D_SISSO = float(input("SISSO Debye temperature θ_D [K]: "))
    T_MIN_FIT     = 5.0
    T_MAX_FIT     = 1000.0

    # Column-width selection
    print("\nFigure width:")
    print("  1 = single-column (89 mm / 3.50 in)  [default]")
    print("  2 = double-column (183 mm / 7.20 in)")
    col_choice = input("  Choice [1]: ").strip() or "1"
    fig_width  = NATURE_DOUBLE_COL_W if col_choice == "2" else NATURE_SINGLE_COL_W
    fig_height = fig_width * (5.0 / 6.5)   # preserve original aspect ratio
    fig_height = min(fig_height, NATURE_MAX_HEIGHT)

    cwd        = os.getcwd()
    input_dir  = prompt_dir("Input directory (PRIMCELL.vasp + thermal_properties.yaml)", default=cwd)
    output_dir = prompt_dir("Output directory", default=input_dir)
    os.makedirs(output_dir, exist_ok=True)

    paths = {
        'PRIMCELL': os.path.join(input_dir, 'PRIMCELL.vasp'),
        'YAML'    : os.path.join(input_dir, 'thermal_properties.yaml'),
    }
    missing = [k for k, p in paths.items() if not os.path.isfile(p)]
    if missing:
        print("\n  ✗ Missing required file(s):")
        for k in missing:
            print(f"      – {paths[k]}")
        sys.exit(1)

    for k, p in paths.items():
        print(f"  ✓  {p}")

    # ── Formula from PRIMCELL.vasp ────────────────────────────────────────
    print("\nReading formula from PRIMCELL.vasp …")
    prim_species              = parse_vasp_species(paths['PRIMCELL'])
    formula_short, formula_latex = build_formula_strings(prim_species)
    print(f"  ✓ Species  : { dict(prim_species) }")
    print(f"  ✓ Formula  : {formula_short}  /  {formula_latex}")

    # ── Load phonopy Cv data ──────────────────────────────────────────────
    print("\nLoading thermal_properties.yaml …")
    T_ph, Cv_ph, N_ATOMS = load_cv_from_yaml(paths['YAML'])
    if N_ATOMS == 0:
        print("  ⚠  'natom' not found in YAML; falling back to PRIMCELL atom count.")
        N_ATOMS = sum(prim_species.values())
    print(f"  ✓ {len(T_ph)} temperature points")
    print(f"  ✓ N_ATOMS (from YAML) : {N_ATOMS}")
    print(f"  ✓ SISSO θ_D           : {THETA_D_SISSO} K")
    print(f"  ✓ Fit range           : {T_MIN_FIT} – {T_MAX_FIT} K")

    Cv_DP = 3.0 * N_ATOMS * R   # Dulong-Petit limit

    # ── Fit Debye model ───────────────────────────────────────────────────
    mask   = (T_ph >= T_MIN_FIT) & (T_ph <= T_MAX_FIT) & (T_ph > 0)
    T_fit  = T_ph[mask]
    Cv_fit = Cv_ph[mask]

    print(f"\nFitting Debye model with {len(T_fit)} points …")
    popt, pcov = curve_fit(
        lambda T, thetaD: debye_Cv_array(T, thetaD, N_ATOMS),
        T_fit, Cv_fit, p0=[500.0], maxfev=20000
    )
    theta_fitted = float(popt[0])
    sigma_theta  = float(np.sqrt(np.diag(pcov))[0])

    Cv_pred = debye_Cv_array(T_fit, theta_fitted, N_ATOMS)
    ss_res  = np.sum((Cv_fit - Cv_pred)**2)
    ss_tot  = np.sum((Cv_fit - np.mean(Cv_fit))**2)
    R2      = 1.0 - ss_res / ss_tot

    delta = theta_fitted - THETA_D_SISSO
    pct   = 100.0 * delta / THETA_D_SISSO

    print(f"\n  {'─'*38}")
    print(f"  Fitted θ_D = {theta_fitted:.3f} ± {sigma_theta:.3f} K")
    print(f"  SISSO  θ_D = {THETA_D_SISSO:.3f} K")
    print(f"  Δ          = {delta:+.2f} K  ({pct:+.2f} %)")
    print(f"  R²         = {R2:.6f}")
    print(f"  {'─'*38}")

    # ── Smooth model curves ───────────────────────────────────────────────
    T_model         = np.linspace(1, T_MAX_FIT, 600)
    Cv_debye_fitted = debye_Cv_array(T_model, theta_fitted,  N_ATOMS)
    Cv_debye_SISSO  = debye_Cv_array(T_model, THETA_D_SISSO, N_ATOMS)

    # ── Output paths ──────────────────────────────────────────────────────
    out_pdf  = os.path.join(output_dir, f"Cv_comparison_{formula_short}.pdf")
    out_tiff = os.path.join(output_dir, f"Cv_comparison_{formula_short}.tiff")

    # ─────────────────────────────────────────────────────────────────────
    # PLOT
    # ─────────────────────────────────────────────────────────────────────
    BOX_LW = 0.75   # spine linewidth (kept in sync with axes.linewidth)

    # Use exact Nature column width; height scaled to preserve original ratio.
    fig, ax = plt.subplots(figsize=(fig_width, fig_height))

    # Margins expressed as fractions of the figure size so that the absolute
    # white-space around the axes is proportional at different column widths.
    fig.subplots_adjust(left=0.13, right=0.97, top=0.97, bottom=0.12)

    # ── Data series (colours / styles unchanged from original) ───────────

    # Phonopy data points
    ax.plot(
        T_ph, Cv_ph,
        'o', color='#5A87E8', markersize=4, zorder=3,
        label=fr"Phonopy  ({formula_latex})",
    )

    # Fitted Debye curve
    ax.plot(
        T_model, Cv_debye_fitted,
        '-', color='#E06633', linewidth=1.8, zorder=4,
        label=(
            fr"Debye fit  "
            fr"($\Theta_D^{{\mathrm{{phonopy}}}}$ = {theta_fitted:.2f}"
            fr" $\pm$ {sigma_theta:.2f} K)"
        ),
    )
    ax.plot(
        T_model, Cv_debye_fitted,
        '-', color='none', linewidth=0.1, zorder=4,
        label=(
            fr"Debye fit "
            fr"($R^2$ = {R2:.4f})"
        ),
    )

    # SISSO Debye curve
    ax.plot(
        T_model, Cv_debye_SISSO,
        '--', color='red', linewidth=1.8, zorder=2,
        label=fr"Debye SISSO  ($\Theta_D^{{\mathrm{{SISSO}}}}$ = {THETA_D_SISSO:.2f} K)",
    )
    ax.plot(
        T_model, Cv_debye_SISSO,
        '--', color='none', linewidth=0.1, zorder=2,
        label=fr"$\Delta$ $\Theta_D$  = {delta:+.2f} K",
    )

    # Dulong-Petit limit
    ax.axhline(
        y=Cv_DP,
        color='#696969', linewidth=1.3, linestyle='--', alpha=1.0,
        label=fr"Dulong–Petit  (3$nR$ = {Cv_DP:.2f} J/mol$\cdot$K)",
        zorder=1,
    )

    # ── Axes labels ───────────────────────────────────────────────────────
    # Font sizes controlled by rcParams (7 pt for labels, 6 pt for ticks);
    # explicit fontsize args kept to ensure they override any theme.
    ax.set_xlabel(r"Temperature (K)", fontsize=7, labelpad=4)
    ax.set_ylabel(r"$C_v$  (J/mol$\cdot$K)", fontsize=7, labelpad=4)

    # ── Tick styling ──────────────────────────────────────────────────────
    ax.tick_params(axis='both', which='major',
                   direction='in', width=0.75, length=3.5,
                   labelsize=6, top=True, right=True)
    ax.tick_params(axis='both', which='minor',
                   direction='in', width=0.5, length=2.0,
                   top=True, right=True)
    ax.yaxis.set_minor_locator(ticker.AutoMinorLocator(5))
    ax.xaxis.set_minor_locator(ticker.AutoMinorLocator(10))

    for sp in ax.spines.values():
        sp.set_linewidth(BOX_LW)

    ax.set_xlim(0, T_MAX_FIT + 100)
    ax.set_ylim(0, Cv_DP * 1.12)

    # ── Legend ────────────────────────────────────────────────────────────
    ax.legend(
        loc='lower right',
        fontsize=6,             # Nature: legend text ≥ 5 pt
        frameon=True,
        framealpha=0.55,
        edgecolor='black',
        fancybox=False,
        handlelength=1.6,
        handleheight=0.9,
        handletextpad=0.5,
        borderpad=0.6,
    )

    # ── Save ─────────────────────────────────────────────────────────────
    print("\nSaving …")

    # PDF – vector format for Nature online submission
    # bbox_inches=None so the figure dimensions are exactly as specified;
    # use fig.savefig with the pre-set subplots_adjust instead.
    fig.savefig(out_pdf, dpi=600, bbox_inches='tight',
                format='pdf', transparent=False)
    print(f"  ✓ {out_pdf}")

    # TIFF – 600 dpi, LZW compression (Nature production standard)
    # Requires Pillow: pip install Pillow
    fig.savefig(out_tiff, dpi=600, bbox_inches='tight',
                format='tiff',
                pil_kwargs={'compression': 'tiff_lzw'})
    print(f"  ✓ {out_tiff}  (600 dpi, LZW)")

    plt.close(fig)

    # ─────────────────────────────────────────────────────────────────────
    # EXCEL source data (unchanged)
    # ─────────────────────────────────────────────────────────────────────
    out_xlsx = os.path.join(output_dir, f"Cv_comparison_{formula_short}_source_data.xlsx")
    wb = openpyxl.Workbook()

    hdr_font   = Font(name='Arial', bold=True, size=10)
    hdr_fill   = PatternFill('solid', start_color='D9E1F2')
    info_fill  = PatternFill('solid', start_color='F2F2F2')
    center     = Alignment(horizontal='center')
    thin       = Side(style='thin')
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header_row(ws, row, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = center
            cell.border    = border

    def write_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # Sheet 1 – Phonopy raw data
    ws1 = wb.active
    ws1.title = 'Phonopy_Data'
    ws1.append(['Temperature (K)', f'Cv_phonopy (J/mol·K)  [{formula_short}]'])
    style_header_row(ws1, 1, 2)
    for t, cv in zip(T_ph, Cv_ph):
        ws1.append([round(float(t), 4), round(float(cv), 6)])
    write_col_widths(ws1, [20, 32])

    # Sheet 2 – Model curves (smooth, 600-point)
    ws2 = wb.create_sheet('Model_Curves')
    dp_col = [Cv_DP] * len(T_model)
    ws2.append([
        'Temperature (K)',
        f'Cv_Debye_fitted (J/mol·K)  [θ_D={theta_fitted:.3f} K]',
        f'Cv_Debye_SISSO (J/mol·K)  [θ_D={THETA_D_SISSO:.3f} K]',
        f'Cv_Dulong_Petit (J/mol·K)  [3nR={Cv_DP:.4f}]',
    ])
    style_header_row(ws2, 1, 4)
    for t, cf, cs, dp in zip(T_model, Cv_debye_fitted, Cv_debye_SISSO, dp_col):
        ws2.append([round(float(t), 4), round(float(cf), 6),
                    round(float(cs), 6), round(float(dp), 6)])
    write_col_widths(ws2, [20, 42, 40, 38])

    # Sheet 3 – Fit summary
    ws3 = wb.create_sheet('Fit_Summary')
    summary_rows = [
        ('Parameter',              'Value',                   'Unit'),
        ('Formula (primitive)',     formula_short,             '—'),
        ('N_atoms (supercell)',     N_ATOMS,                   'atoms'),
        ('θ_D  (fitted)',           round(theta_fitted, 4),    'K'),
        ('σ(θ_D)  (fitted)',        round(sigma_theta, 4),     'K'),
        ('θ_D  (SISSO)',            THETA_D_SISSO,             'K'),
        ('Δθ_D  (fitted − SISSO)', round(delta, 4),           'K'),
        ('Δθ_D  (%)',               round(pct, 4),             '%'),
        ('R²',                      round(R2, 8),              '—'),
        ('Fit T_min',               T_MIN_FIT,                 'K'),
        ('Fit T_max',               T_MAX_FIT,                 'K'),
        ('Dulong–Petit limit 3nR',  round(Cv_DP, 6),          'J/mol·K'),
    ]
    for i, row in enumerate(summary_rows, 1):
        ws3.append(list(row))
        if i == 1:
            style_header_row(ws3, 1, 3)
        else:
            for c in range(1, 4):
                cell = ws3.cell(row=i, column=c)
                cell.fill   = info_fill
                cell.border = border
                cell.font   = Font(name='Arial', size=10)
    write_col_widths(ws3, [28, 22, 14])

    wb.save(out_xlsx)
    print(f"  ✓ {out_xlsx}")
    print("\nDone. ✓")
    print("=" * 62)


if __name__ == "__main__":
    main()
