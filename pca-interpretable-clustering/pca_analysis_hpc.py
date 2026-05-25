"""
================================================================================
PCA ANALYSIS OF MATERIALS DATASET — HPC VERSION
================================================================================
Usage:
    python pca_analysis_hpc.py            # reads Material_classification_PCA.txt
    python pca_analysis_hpc.py --no-log   # skip writing a .log file
================================================================================
"""

# ── Standard library ──────────────────────────────────────────────────────────
import os
import sys
import time
import datetime
import argparse
import logging
import platform
import warnings
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import io
import codecs

warnings.filterwarnings('ignore')

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')   # non-interactive backend — required for HPC
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.patches as patches
from matplotlib.colors import LinearSegmentedColormap
from matplotlib.lines import Line2D
from matplotlib.ticker import AutoMinorLocator, MaxNLocator
from matplotlib.legend_handler import HandlerBase
from matplotlib.patches import Patch
import matplotlib.colors as mcolors
from matplotlib.collections import PolyCollection

from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler
from scipy.interpolate import griddata, make_interp_spline

from joblib import Parallel, delayed, cpu_count as joblib_cpu_count
from concurrent.futures import ProcessPoolExecutor, as_completed

import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
import seaborn as sns

try:
    import psutil
except ImportError:
    psutil = None

try:
    import plotly.graph_objects as go
    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False

plt.rcParams['font.family'] = 'serif'
plt.rcParams['font.serif']  = ['DejaVu Serif', 'Times New Roman', 'Palatino']
sns.set_style("ticks")

# ============================================================================
# LOGGING  (replaces Colors terminal output — works cleanly in HPC job logs)
# ============================================================================
def setup_logging(log_path: str = None) -> logging.Logger:
    logger = logging.getLogger("PCA_HPC")
    logger.handlers.clear()          
    logger.propagate = False         
    logger.setLevel(logging.DEBUG)
    fmt = logging.Formatter("%(asctime)s  %(levelname)-8s  %(message)s",
                            datefmt="%Y-%m-%d %H:%M:%S")
    try:
        utf8_stdout = open(sys.stdout.fileno(), mode='w', encoding='utf-8', buffering=1)
    except (io.UnsupportedOperation, AttributeError):
        utf8_stdout = codecs.getwriter('utf-8')(sys.stdout.buffer) \
            if hasattr(sys.stdout, 'buffer') else sys.stdout

    ch = logging.StreamHandler(utf8_stdout)
    ch.setFormatter(fmt)
    logger.addHandler(ch)

    if log_path:
        fh = logging.FileHandler(log_path, encoding='utf-8')
        fh.setFormatter(fmt)
        logger.addHandler(fh)
    return logger

log: logging.Logger = None   # set in main()

# ============================================================================
# CONFIG LOADER  — reads PCA_config.txt from same folder
# ============================================================================
def load_config() -> dict:
    folder   = os.path.dirname(os.path.abspath(__file__)) if "__file__" in dir() else os.getcwd()
    cfg_path = os.path.join(folder, "PCA_config.txt")
    if not os.path.isfile(cfg_path):
        raise FileNotFoundError(f"PCA_config.txt not found in: {folder}")

    raw = {}
    with open(cfg_path, "r", encoding="utf-8") as f:
        for line in f:
            stripped = line.rstrip()
            if not stripped or stripped.lstrip().startswith("#"):
                continue
            if "=" in stripped:
                key, _, value = stripped.partition("=")
                raw[key.strip()] = value.strip()

    def _str(k, default=None):
        v = raw.get(k, "").strip()
        return v if v else default

    def _bool(k, default=False):
        v = (_str(k) or "").lower()
        return v in ("yes", "y", "true", "1")

    def _int(k, default=1):
        v = _str(k)
        try:
            return int(v) if v else default
        except ValueError:
            return default

    def _list_str(k):
        v = _str(k, "")
        return [x.strip() for x in v.split(",") if x.strip()] if v else []

    return {
        "data_folder":        _str("data_folder"),
        "file_subset":        _list_str("file_subset"),   # [] = load all
        "feature_columns":    _list_str("feature_columns"),  # [] = use all
        "n_cpus":             _int("n_cpus", 1),
        "save_dir":           _str("save_dir"),
        "save_name":          _str("save_name", "pca_results"),
        "symbol_file":        _str("symbol_file"),
        "interactive_scatter":_bool("interactive_scatter"),
        "inject_file":        _str("inject_file"),
        "symbol_file_html":   _str("symbol_file_html"),
        "show_labels":        _bool("show_labels"),
    }

# ============================================================================
# KNOWN UNITS PER COLUMN KEY
# ============================================================================
COLUMN_UNITS: Dict[str, str] = {
    'NE': '', 'NA': '', 'BGAP': 'eV', 'EPA': 'eV/atom', 'FEPA': 'eV/atom',
    'EFMR': 'eV', 'TM': r'μ_B', 'MSITE': '', 'ASPIN': r'μ_B',
    'V': 'Å³', 'VPA': 'Å³/atom', 'D': 'g/cm³', 'PR': '', 'ANI': '',
    'BM': 'GPa', 'SM': 'GPa', 'YM': 'GPa', 'PUGR': '', 'IPUGR': '',
    'ENPA': 'eV/atom', 'vm': 'm/s', 'vl': 'm/s', 'vs': 'm/s',
    'DT_AGL': 'K', 'DT_A_AGL': 'K', 'DT': 'K',
    'Cp_300': 'J/(mol·K)', 'Cv_300': 'J/(mol·K)', 'TC_300': 'W/(m·K)',
    'TEX_300': '1/K', 'VIB_EN_300': 'eV/atom', 'VIB_FE_300': 'eV/atom',
    'gruneisen': '',
}

# ============================================================================
# FIGURES HELPER CLASSES
# ============================================================================
class TwoColorLine:
    def __init__(self, color1, color2, linestyle='-', linewidth=2):
        self.color1 = color1; self.color2 = color2
        self.linestyle = linestyle; self.linewidth = linewidth


class HandlerTwoColorLine(HandlerBase):
    def __init__(self, gap_ratio=0.05, **kwargs):
        self.gap_ratio = gap_ratio
        super().__init__(**kwargs)

    def create_artists(self, legend, orig_handle, xdescent, ydescent,
                       width, height, fontsize, trans):
        gap  = width * self.gap_ratio
        half = (width - gap) / 2
        line1 = Line2D([xdescent, xdescent + half],
                       [ydescent + height / 2] * 2,
                       linestyle=orig_handle.linestyle,
                       linewidth=orig_handle.linewidth,
                       color=orig_handle.color1, transform=trans)
        line2 = Line2D([xdescent + half + gap, xdescent + width],
                       [ydescent + height / 2] * 2,
                       linestyle=orig_handle.linestyle,
                       linewidth=orig_handle.linewidth,
                       color=orig_handle.color2, transform=trans)
        return [line1, line2]

# ============================================================================
# ENVIRONMENT & PARALLEL BACKEND
# ============================================================================
def _detect_environment() -> dict:
    in_jupyter = False
    try:
        shell = get_ipython().__class__.__name__   # type: ignore
        if 'ZMQ' in shell or 'Kernel' in shell:
            in_jupyter = True
    except NameError:
        pass

    slurm_cpus   = int(os.environ.get('SLURM_CPUS_PER_TASK', 0))
    in_slurm     = slurm_cpus > 0
    machine_cpus = joblib_cpu_count()

    log.info(f"Environment: {'Jupyter' if in_jupyter else 'Script'} | "
             f"{'SLURM (HPC)' if in_slurm else 'Local'} | "
             f"vCPUs available: {machine_cpus}")
    return {
        'in_jupyter':   in_jupyter,
        'in_slurm':     in_slurm,
        'slurm_cpus':   slurm_cpus,
        'machine_cpus': machine_cpus,
    }


def _select_backend(n_cpus: int, env: dict) -> dict:
    if n_cpus <= 1:
        return {'backend': 'sequential', 'n_jobs': 1,
                'label': 'Sequential (1 core)'}
    if env['in_jupyter']:
        return {'backend': 'joblib_threads', 'n_jobs': n_cpus,
                'label': f'joblib threads ({n_cpus} cores) — Jupyter safe'}
    if env['in_slurm']:
        effective = min(n_cpus, env['slurm_cpus'])
        return {'backend': 'processpool', 'n_jobs': effective,
                'label': f'ProcessPoolExecutor ({effective} cores) — HPC/SLURM'}
    return {'backend': 'joblib_loky', 'n_jobs': n_cpus,
            'label': f'joblib loky ({n_cpus} cores) — local script'}


def run_parallel(fn, args_list: list, backend_cfg: dict) -> list:
    backend = backend_cfg['backend']
    n_jobs  = backend_cfg['n_jobs']

    if backend == 'sequential' or len(args_list) <= 1:
        return [fn(args) for args in args_list]
    if backend == 'joblib_threads':
        return Parallel(n_jobs=n_jobs, prefer='threads', verbose=0)(
            delayed(fn)(args) for args in args_list)
    if backend == 'joblib_loky':
        return Parallel(n_jobs=n_jobs, backend='loky', verbose=0)(
            delayed(fn)(args) for args in args_list)
    if backend == 'processpool':
        results = [None] * len(args_list)
        def _indexed(idx_args):
            idx, args = idx_args
            return idx, fn(args)
        with ProcessPoolExecutor(max_workers=n_jobs) as ex:
            futures = {ex.submit(_indexed, (i, a)): i
                       for i, a in enumerate(args_list)}
            for fut in as_completed(futures):
                idx, result = fut.result()
                results[idx] = result
        return results
    raise ValueError(f"Unknown backend: {backend}")

# ============================================================================
# SYMBOL CONVERTER
# ============================================================================
class SymbolConverter:
    def __init__(self, filepath: Optional[str] = None):
        self.mapping: Dict[str, str] = {}
        if filepath:
            # Accept folder or direct file path
            if os.path.isdir(filepath):
                candidate = os.path.join(filepath, "symbol_conversion.txt")
                filepath = candidate if os.path.isfile(candidate) else None
            if filepath and os.path.isfile(filepath):
                self._load_from_file(filepath)
            else:
                log.warning("Symbol file not found — using defaults.")
                self._load_defaults()
        else:
            self._load_defaults()

    def _load_from_file(self, filepath: str) -> None:
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith('#'):
                        continue
                    if '=' in line and '->' not in line:
                        key, value = line.split('=', 1)
                    elif '->' in line:
                        key, value = line.split('->', 1)
                    else:
                        continue
                    key = key.strip(); value = value.strip()
                    if key and value:
                        self.mapping[key] = value
            log.info(f"Loaded {len(self.mapping)} symbol mappings from {filepath}")
        except Exception as e:
            log.warning(f"Failed to load symbol file: {e}. Using defaults.")
            self._load_defaults()

    def _load_defaults(self) -> None:
        self.mapping = {
            'SM': 'G_vrh', 'YM': 'E', 'VPA': 'V_atom',
            'D': 'rho', 'DT': 'Theta_D',
        }
        log.info("Using default symbol mappings")

    def get(self, key: str, default: Optional[str] = None) -> str:
        return self.mapping.get(key, default if default is not None else key)

    def get_axis_label(self, key: str) -> str:
        symbol = self.get(key)
        unit   = COLUMN_UNITS.get(key, '')
        return f"{symbol} [{unit}]" if unit else symbol

    def labels_for_columns(self, columns: List[str]) -> List[str]:
        return [self.get_axis_label(col) for col in columns]

    def report_coverage(self, columns: List[str]) -> None:
        mapped   = [c for c in columns if c in self.mapping]
        unmapped = [c for c in columns if c not in self.mapping]
        if mapped:   log.info(f"Symbol mappings found for: {mapped}")
        if unmapped: log.warning(f"No symbol mapping for: {unmapped} — using column name")

# ============================================================================
# CATEGORY CONFIGURATION
# ============================================================================
class CategoryConfig:
    FILENAME_MAP: Dict[str, str] = {
        "element": "Single Element",
        "metal":   "Metal & Metalloid Compounds",
        "boride":  "Borides",
        "carbide": "Carbides",
        "oxide":   "Oxides",
        "nitride": "Nitrides",
        "halide":  "Halides",
        "other":   "Other Compounds",
    }
    DISPLAY_ORDER: List[str] = [
        "Single Element", "Metal & Metalloid Compounds",
        "Nitrides", "Other Compounds", "Oxides",
        "Halides", "Borides", "Carbides",
    ]
    COLORS: Dict[str, str] = {
        "Single Element":              'teal',
        "Metal & Metalloid Compounds": 'tab:orange',
        "Oxides":                      '#FFFF33',
        "Nitrides":                    'fuchsia',
        "Halides":                     'lime',
        "Borides":                     'deepskyblue',
        "Carbides":                    'crimson',
        "Other Compounds":             'blue',
    }
    MARKERS: Dict[str, str] = {
        "Single Element":              'o',
        "Metal & Metalloid Compounds": 's',
        "Oxides":                      'P',
        "Nitrides":                    '*',
        "Halides":                     'D',
        "Borides":                     'X',
        "Carbides":                    '^',
        "Other Compounds":             'v',
    }
    MARKER_SIZES: Dict[str, int] = {
        'o': 95, 's': 65, 'D': 48, '^': 115,
        '*': 225, 'P': 115, 'X': 105, 'v': 75,
    }
    XL_HEADER_FILLS: Dict[str, str] = {
        "Single Element":              "008080",
        "Metal & Metalloid Compounds": "FFA500",
        "Oxides":                      "FFFF33",
        "Nitrides":                    "FF00FF",
        "Halides":                     "00FF00",
        "Borides":                     "00BFFF",
        "Carbides":                    "DC143C",
        "Other Compounds":             "0000FF",
    }

    @classmethod
    def categorize(cls, filename: str) -> str:
        name = os.path.basename(filename).lower()
        for keyword, label in cls.FILENAME_MAP.items():
            if keyword in name:
                return label
        return "Other Compounds"

# ============================================================================
# DATA LOADER
# ============================================================================
def _load_single(args):
    folder, file, required, categorize_fn = args
    full_path = os.path.join(folder, file)
    try:
        df = pd.read_excel(full_path, engine='openpyxl')
        extra_id_cols = [c for c in ['material_id', 'crystal_system'] if c in df.columns]
        available = extra_id_cols + [c for c in required if c in df.columns]
        missing   = [c for c in required if c not in df.columns]
        if missing:
            return None, file, f"missing columns {missing}"
        df = df[available].dropna()
        df["Category"] = categorize_fn(full_path)
        return df, file, None
    except Exception as e:
        return None, file, str(e)


class DataLoader:
    REQUIRED_COLUMNS = ["formula", "SM", "YM", "VPA", "D", "DT"]
    FEATURE_COLUMNS  = ["SM", "YM", "VPA", "D", "DT"]

    @staticmethod
    def load_folder(folder: str,
                    file_subset: Optional[List[str]] = None,
                    feature_cols: Optional[List[str]] = None,
                    n_cpus: int = 1,
                    backend_cfg: dict = None) -> pd.DataFrame:

        log.info("Loading data ...")
        xlsx_files = sorted([f for f in os.listdir(folder) if f.endswith(".xlsx")])
        if not xlsx_files:
            log.error(f"No .xlsx files found in: {folder}"); sys.exit(1)

        # Apply file subset from config
        if file_subset:
            resolved = []
            for tok in file_subset:
                if tok.isdigit():
                    idx = int(tok) - 1
                    if 0 <= idx < len(xlsx_files):
                        resolved.append(xlsx_files[idx])
                    else:
                        log.warning(f"File index {tok} out of range — skipped.")
                else:
                    if tok in xlsx_files:
                        resolved.append(tok)
                    else:
                        log.warning(f"File '{tok}' not found — skipped.")
            xlsx_files = resolved if resolved else xlsx_files

        # Use feature_cols to build required columns list
        required = ["formula"] + (feature_cols if feature_cols else DataLoader.FEATURE_COLUMNS)

        args_list = [(folder, f, required, CategoryConfig.categorize) for f in xlsx_files]
        backend_cfg = backend_cfg or {'backend': 'sequential', 'n_jobs': 1, 'label': 'Sequential'}
        log.info(f"Loading {len(xlsx_files)} file(s) with: {backend_cfg['label']}")

        results = run_parallel(_load_single, args_list, backend_cfg)

        frames = []
        for df_loaded, fname, err in results:
            if err:
                log.warning(f"{fname} — {err}, skipping.")
            else:
                log.info(f"{fname}  ->  '{df_loaded['Category'].iloc[0]}'  ({len(df_loaded):,} rows)")
                frames.append(df_loaded)

        if not frames:
            log.error("No valid data loaded."); sys.exit(1)

        combined = pd.concat(frames, ignore_index=True)
        log.info(f"Total rows loaded: {len(combined):,}")
        return combined

    @staticmethod
    def resolve_feature_columns(df: pd.DataFrame,
                                 feature_subset: List[str]) -> List[str]:
        """
        Resolve feature_columns from config:
        - empty list  -> use all FEATURE_COLUMNS present in df
        - list of ints (as strings) -> serial number selection from available
        - list of names -> direct column names
        """
        available = [c for c in DataLoader.FEATURE_COLUMNS if c in df.columns]
        if not feature_subset:
            return available

        resolved = []
        for tok in feature_subset:
            if tok.isdigit():
                idx = int(tok) - 1
                if 0 <= idx < len(available):
                    resolved.append(available[idx])
                else:
                    log.warning(f"Feature index {tok} out of range — skipped.")
            else:
                if tok in available:
                    resolved.append(tok)
                else:
                    log.warning(f"Feature column '{tok}' not found — skipped.")
        return resolved if resolved else available

# ============================================================================
# PCA ALGORITHM
# ============================================================================
class PCARunner:
    def __init__(self, n_components: int = 5):
        self.n_components = n_components
        self.scaler = StandardScaler()
        self.pca    = PCA(n_components=n_components, svd_solver='randomized', random_state=42)

    def fit_transform(self, df: pd.DataFrame,
                      feature_cols: List[str]) -> Tuple[pd.DataFrame, np.ndarray]:
        log.info("BLOCK 1: PCA — fitting scaler and decomposing ...")
        _t0 = time.perf_counter()
        X_scaled = self.scaler.fit_transform(df[feature_cols])
        log.info(f"  Scaler: {time.perf_counter()-_t0:.3f}s")
        _t0 = time.perf_counter()
        X_pca    = self.pca.fit_transform(X_scaled)
        log.info(f"  PCA decomposition: {time.perf_counter()-_t0:.3f}s")

        pc_cols = [f'PC{i+1}' for i in range(X_pca.shape[1])]
        pca_df  = pd.DataFrame(X_pca, columns=pc_cols)
        result  = pd.concat([df.reset_index(drop=True), pca_df], axis=1)

        evr = self.pca.explained_variance_ratio_
        cum = np.cumsum(evr)
        log.info(f"  {'Component':<12} {'Variance':>10} {'Cumulative':>12}")
        for i, (v, c) in enumerate(zip(evr, cum)):
            log.info(f"  PC{i+1:<9}  {v:>9.4f}   {c:>10.4f}")

        loadings = self.pca.components_.T
        log.info("PCA complete")
        return result, loadings

    @property
    def explained_variance(self) -> np.ndarray:
        return self.pca.explained_variance_ratio_


# ============================================================================
# EXCEL HELPER FUNCTIONS  (identical to source)
# ============================================================================
_THIN_BORDER = Border(
    left=Side(style='thin'),  right=Side(style='thin'),
    top=Side(style='thin'),   bottom=Side(style='thin'),
)

def _xl_header_style(cell, hex_fill="1F4E79", font_hex="FFFFFF",
                     bold=True, font_size=11):
    cell.font      = Font(bold=bold, color=font_hex, name='Arial', size=font_size)
    cell.fill      = PatternFill("solid", fgColor=hex_fill)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border    = _THIN_BORDER

def _xl_data_style(cell, number_fmt=None, align='center'):
    cell.font      = Font(name='Arial', size=10)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border    = _THIN_BORDER
    if number_fmt:
        cell.number_format = number_fmt

def _xl_auto_width(ws, min_w=10, max_w=40):
    for col_cells in ws.columns:
        length = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = \
            max(min_w, min(length + 3, max_w))

def _write_df_to_sheet(ws, df: pd.DataFrame, header_fill="1F4E79",
                       header_font="FFFFFF", num_fmt='0.0000', freeze=True):
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_name))
        _xl_header_style(cell, hex_fill=header_fill, font_hex=header_font)
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            fmt  = num_fmt if isinstance(value, (float, np.floating)) else None
            _xl_data_style(cell, number_fmt=fmt)
    if freeze:
        ws.freeze_panes = ws.cell(row=2, column=1)
    _xl_auto_width(ws)

def _compute_pc_bin_stats(df, pc_col, dt_col, bin_size):
    bins = np.arange(df[pc_col].min(), df[pc_col].max() + bin_size, bin_size)
    grp  = df.groupby(pd.cut(df[pc_col], bins=bins)).agg(
        mean_PC   = (pc_col, 'mean'), count     = (dt_col, 'size'),
        DT_mean   = (dt_col, 'mean'), DT_std    = (dt_col, 'std'),
        DT_min    = (dt_col, 'min'),  DT_max    = (dt_col, 'max'),
        DT_p10    = (dt_col, lambda x: np.percentile(x, 10) if len(x) else np.nan),
        DT_p90    = (dt_col, lambda x: np.percentile(x, 90) if len(x) else np.nan),
    ).reset_index(drop=True)
    grp.rename(columns={'mean_PC': pc_col}, inplace=True)
    return grp

def _compute_dt_distribution(df, dt_col, bin_size=75):
    bins = np.arange(df[dt_col].min(), df[dt_col].max() + bin_size, bin_size)
    df2  = df.copy()
    df2['bin_centre'] = pd.cut(df2[dt_col], bins=bins, labels=bins[:-1]).astype(float)
    grp  = df2.groupby('bin_centre').size().reset_index(name='count')
    grp.rename(columns={'bin_centre': f'{dt_col}_bin_centre (K)'}, inplace=True)
    return grp

def _compute_threshold_ranges(df, feature_cols, dt_col, low_cut=350, high_cut=1000):
    dt_min   = df[dt_col].min(); dt_max = df[dt_col].max()
    low_thr  = np.arange(max(-1000, dt_min), low_cut + 1, 2)
    high_thr = np.arange(high_cut, dt_max + 1, 2)

    def build_range_df(thresholds, mask_fn):
        rows = []
        for thr in thresholds:
            sub = df[mask_fn(thr)]
            if not len(sub): continue
            row = {'Threshold_K': thr, 'N_materials': len(sub)}
            for feat in feature_cols:
                if feat in sub.columns:
                    row[f'{feat}_min']  = sub[feat].min()
                    row[f'{feat}_max']  = sub[feat].max()
                    row[f'{feat}_mean'] = sub[feat].mean()
            rows.append(row)
        return pd.DataFrame(rows)

    return (build_range_df(low_thr,  lambda t: df[dt_col] < t),
            build_range_df(high_thr, lambda t: df[dt_col] > t))


# ============================================================================
# PLOT FUNCTIONS  (identical logic to source; Colors.* -> log.*)
# ============================================================================
def plot_explained_variance(explained_variance, save_base):
    mask   = explained_variance > 1e-3
    ev     = explained_variance[mask]
    n_bars = len(ev)
    gradient_cmap = LinearSegmentedColormap.from_list(
        '3d_blue', ['#0000E6', '#0343DF', '#0692E7'], N=2560)
    gradient_img = np.repeat(np.linspace(0, 1, 256).reshape(1, -1), 256, axis=0)

    fig, ax = plt.subplots(figsize=(6, 5))
    bar_width = 0.75
    for i, var in enumerate(ev):
        xc = i + 1
        xl, xr = xc - bar_width/2, xc + bar_width/2
        ax.imshow(gradient_img, aspect='auto',
                  extent=[xl, xr, 0, var], origin='lower', cmap=gradient_cmap,
                  vmin=0, vmax=1, zorder=2)
        rect = mpatches.FancyBboxPatch(
            (xl, 0), bar_width, var, boxstyle="square,pad=0",
            linewidth=1.82, edgecolor='darkblue', facecolor='none', zorder=3)
        ax.add_patch(rect)
        ax.text(xc, var + 0.01, f"{var:.2f}", ha='center', fontsize=16, zorder=4)
    ax.set_xlabel('PCA Principal Components', fontsize=20)
    ax.set_ylabel('Explained Variance', fontsize=20)
    ax.set_xticks(range(1, n_bars + 1))
    ax.tick_params(axis='both', which='both', direction='in',
                   labelsize=18.5, length=8, top=False, bottom=False,
                   left=True, right=True)
    ax.tick_params(axis='both', which='minor', direction='in',
                   length=3.5, width=1.0, top=False, bottom=False,
                   left=True, right=True)
    ax.yaxis.set_major_locator(MaxNLocator(nbins=6))
    ax.minorticks_on()
    ax.xaxis.set_minor_locator(AutoMinorLocator(10))
    ax.yaxis.set_minor_locator(AutoMinorLocator(10))
    ax.set_xlim(0.5, n_bars + 0.5)
    ax.set_ylim(0, max(ev) * 1.25)
    plt.tight_layout()
    for fmt in ['png', 'pdf']:
        plt.savefig(f"{save_base}_Explained_Variance.{fmt}", dpi=600, bbox_inches="tight")
    plt.close()
    log.info(f"Variance plot saved: {save_base}_Explained_Variance.png/pdf")


def plot_loadings(loadings, feature_labels, save_base):
    MARKERS = ['o', 's', 'D', '^', 'H']
    COLORS  = ['red', 'blue', 'green', 'orange', 'purple']
    fig, ax = plt.subplots(figsize=(8.5, 5.7))
    for i, label in enumerate(feature_labels):
        ax.scatter(loadings[i, 0], loadings[i, 1],
                   marker=MARKERS[i % len(MARKERS)], color=COLORS[i % len(COLORS)],
                   s=50, label=label, edgecolors='black', zorder=3)
    ax.axhline(0, color='black', linewidth=1, linestyle='--')
    ax.axvline(0, color='black', linewidth=1, linestyle='--')
    ax.add_patch(patches.Ellipse((0, 0), width=2, height=2,
                                  fill=False, color='black', linestyle='-', linewidth=1))
    ax.set_xlabel("PC1 Loading", fontsize=18); ax.set_ylabel("PC2 Loading", fontsize=18)
    ax.set_xlim(-1, 1); ax.set_ylim(-1, 1)
    ax.tick_params(axis='both', which='both', direction='in', labelsize=15, length=8,
                   top=True, bottom=True, left=True, right=True)
    ax.tick_params(axis='both', which='minor', direction='in', length=3.5, width=1.0,
                   top=True, bottom=True, left=True, right=True)
    ax.xaxis.set_major_locator(MaxNLocator(nbins=6))
    ax.yaxis.set_major_locator(MaxNLocator(nbins=6))
    ax.minorticks_on()
    ax.xaxis.set_minor_locator(AutoMinorLocator(10))
    ax.yaxis.set_minor_locator(AutoMinorLocator(7))
    ax.legend(loc=[1.04, 0.42], fontsize=15, frameon=True, edgecolor='black')
    plt.tight_layout()
    for fmt in ['png', 'pdf']:
        plt.savefig(f"{save_base}_PC_Loadings.{fmt}", dpi=600, bbox_inches="tight")
    plt.close()
    log.info(f"Loading plot saved: {save_base}_PC_Loadings.png/pdf")


def plot_scatter(df, save_base, show_labels):
    fig, ax = plt.subplots(figsize=(18, 8))
    cfg = CategoryConfig
    for category in cfg.DISPLAY_ORDER:
        subset = df[df["Category"] == category]
        if subset.empty: continue
        marker = cfg.MARKERS[category]
        color  = cfg.COLORS[category]
        size   = cfg.MARKER_SIZES.get(marker, 65)
        ax.scatter(subset['PC1'], subset['PC2'], color=color, marker=marker,
                   edgecolor='black', s=size, alpha=0.7, zorder=2)
        if show_labels:
            for _, row in subset.iterrows():
                ax.annotate(row['formula'], (row['PC1'], row['PC2']),
                            textcoords="offset points", xytext=(3, 3),
                            ha='left', fontsize=6, alpha=0.75)
    ax.axhline(0, color='black', linewidth=1.5, linestyle='--')
    ax.axvline(0, color='black', linewidth=1.5, linestyle='--')
    ax.set_xlabel('Principal Component 1', fontsize=25)
    ax.set_ylabel('Principal Component 2', fontsize=25)
    ax.tick_params(axis='both', which='major', direction='in', labelsize=22, length=16,
                   top=True, bottom=True, left=True, right=True)
    ax.tick_params(axis='both', which='minor', direction='in', length=7, width=1.0,
                   top=True, bottom=True, left=True, right=True)
    ax.yaxis.set_major_locator(MaxNLocator(5))
    ax.minorticks_on()
    ax.xaxis.set_minor_locator(AutoMinorLocator(20))
    ax.yaxis.set_minor_locator(AutoMinorLocator(10))
    legend_elements = [
        Line2D([0], [0], marker=cfg.MARKERS[cat], color='w', label=cat,
               markerfacecolor=cfg.COLORS[cat], markeredgecolor='black',
               markersize=15, linestyle='None')
        for cat in cfg.DISPLAY_ORDER if not df[df["Category"] == cat].empty
    ]
    ax.legend(handles=legend_elements, loc='best', fontsize=20,
              frameon=True, edgecolor='black')
    plt.tight_layout()
    for fmt in ['png', 'pdf']:
        plt.savefig(f"{save_base}_PCA_Scatter.{fmt}", dpi=600, bbox_inches="tight")
    plt.close()
    log.info(f"Scatter plot saved: {save_base}_PCA_Scatter.png/pdf")


def plot_scatter_interactive(df, feature_cols, symbol_converter, save_base,
                              runner=None, inject_path=None):
    if not PLOTLY_AVAILABLE:
        log.warning("plotly not installed — skipping interactive scatter.")
        return
    cfg = CategoryConfig

    def fmt_label(feat):
        symbol = symbol_converter.get(feat, feat)
        unit   = COLUMN_UNITS.get(feat, '')
        return f"{symbol} ({unit})" if unit else symbol

    MPL_TO_HEX = {
        'tab:orange': '#ff7f0e', 'teal': '#008080', 'fuchsia': '#ff00ff',
        'lime': '#00ff00', 'deepskyblue': '#00bfff', 'crimson': '#dc143c',
        'blue': '#0000ff',
    }
    PLOTLY_MARKERS = {
        'o': 'circle', 's': 'square', 'P': 'cross', '*': 'star',
        'D': 'diamond', 'X': 'x', '^': 'triangle-up', 'v': 'triangle-down',
    }

    fig = go.Figure()
    pc_cols_present = [c for c in df.columns if c.startswith('PC')]

    for category in cfg.DISPLAY_ORDER:
        subset = df[df["Category"] == category].copy()
        if subset.empty: continue
        marker_symbol = cfg.MARKERS.get(category, 'o')
        color         = MPL_TO_HEX.get(cfg.COLORS.get(category, 'blue'), cfg.COLORS.get(category, 'blue'))
        plotly_marker = PLOTLY_MARKERS.get(marker_symbol, 'circle')
        mpl_size      = cfg.MARKER_SIZES.get(marker_symbol, 65)
        plotly_size   = max(9, min(10, int(mpl_size ** 0.52)))

        hover_texts = []
        for _, row in subset.iterrows():
            lines = []
            formula = row.get('formula', '')
            crystal = row.get('crystal_system', '')
            lines.append(f"<b>{formula} ({crystal})</b>" if crystal else f"<b>{formula}</b>")
            lines.append(f"<i>{category}</i>")
            lines.append("─" * 26)
            lines.append("<b>PCA Scores</b>")
            for pc in pc_cols_present:
                if pc in row: lines.append(f"&nbsp;&nbsp;{pc}: {row[pc]:+.4f}")
            lines.append("─" * 26)
            lines.append("<b>Properties</b>")
            for feat in feature_cols:
                if feat not in row: continue
                val = row[feat]
                lines.append(f"&nbsp;&nbsp;{fmt_label(feat)}: {val:.4f}" if isinstance(val, float) else f"&nbsp;&nbsp;{fmt_label(feat)}: {val}")
            hover_texts.append("<br>".join(lines))

        fig.add_trace(go.Scatter(
            x=subset['PC1'], y=subset['PC2'], mode='markers', name=category,
            text=hover_texts, hovertemplate="%{text}<extra></extra>",
            marker=dict(symbol=plotly_marker, size=plotly_size, color=color,
                        line=dict(color='black', width=0.9), opacity=0.85),
        ))

    fig.update_layout(
        title=dict(text="<b>PCA Scatter — PC1 vs PC2</b>", x=0.5, xanchor='center'),
        xaxis=dict(title="Principal Component 1", zeroline=True, zerolinecolor='black',
                   showgrid=True, mirror=True, ticks='inside', showline=True),
        yaxis=dict(title="Principal Component 2", zeroline=True, zerolinecolor='black',
                   showgrid=True, mirror=True, ticks='inside', showline=True),
        legend=dict(title=dict(text="<b>Category</b>"), bgcolor='rgba(255,255,255,0.95)',
                    x=1.02, y=1.0, xanchor='left', yanchor='top'),
        plot_bgcolor='white', paper_bgcolor='#ffffff', hovermode='closest',
        width=1500, height=680, margin=dict(l=80, r=230, t=120, b=90),
    )

    out = f"{save_base}_PCA_Scatter_Interactive.html"
    base_fname = os.path.basename(save_base) + '_PCA_scatter'
    config = dict(displayModeBar=True, displaylogo=False, scrollZoom=True,
                  toImageButtonOptions=dict(format='png', filename=base_fname,
                                            height=680, width=1500, scale=3))
    fig.write_html(out, include_plotlyjs=True, config=config, full_html=True)

    INJECT = ""
    if inject_path and os.path.isfile(inject_path):
        with open(inject_path, 'r', encoding='utf-8') as fh:
            INJECT = fh.read().replace('{base_fname}', base_fname)
    else:
        log.warning("Inject file not found — interactive plot saved without custom UI.")

    with open(out, 'r', encoding='utf-8') as fh:
        html = fh.read()
    with open(out, 'w', encoding='utf-8') as fh:
        fh.write(html.replace('</body>', INJECT + '\n</body>'))

    log.info(f"Interactive scatter saved: {out}")


def plot_material_count_vs_components(df, save_base):
    log.info("Generating Material Count vs PC1/PC2 plot ...")
    sorted_data  = df.sort_values("PC1").reset_index(drop=True)
    bin_size_PC1 = 0.950; bin_size_PC2 = 0.475
    bins_PC1 = np.arange(sorted_data["PC1"].min(), sorted_data["PC1"].max() + bin_size_PC1, bin_size_PC1)
    bins_PC2 = np.arange(sorted_data["PC2"].min(), sorted_data["PC2"].max() + bin_size_PC2, bin_size_PC2)

    def _bin_count(data, pc_col, bins):
        grp = (data.groupby(pd.cut(data[pc_col], bins=bins))
               .agg(mean_pc=(pc_col, "mean"), count=("DT", "size"))
               .reset_index(drop=True))
        x = grp["mean_pc"].values
        y = np.where(grp["count"].values == 0, np.nan, grp["count"].values.astype(float))
        return x, y

    def _smooth_segments(x, y, n_pts=400):
        segs_x, segs_y = [], []
        valid   = ~np.isnan(y)
        indices = np.where(np.diff(valid.astype(int)))[0] + 1
        splits  = np.split(np.arange(len(x)), indices)
        for idx_run in splits:
            if len(idx_run) < 2: continue
            xs, ys = x[idx_run], y[idx_run]
            if np.any(np.isnan(ys)): continue
            k = min(3, len(xs) - 1)
            xnew = np.linspace(xs.min(), xs.max(), n_pts)
            segs_x.append(xnew); segs_y.append(make_interp_spline(xs, ys, k=k)(xnew))
        return segs_x, segs_y

    def _marker_idx(x, n_markers=10):
        return np.linspace(0, len(x) - 1, max(1, min(n_markers, len(x))), dtype=int)

    x_PC1, y_PC1 = _bin_count(sorted_data, "PC1", bins_PC1)
    x_PC2, y_PC2 = _bin_count(sorted_data, "PC2", bins_PC2)
    segs_x_PC1, segs_y_PC1 = _smooth_segments(x_PC1, y_PC1)
    segs_x_PC2, segs_y_PC2 = _smooth_segments(x_PC2, y_PC2)

    fig, ax1 = plt.subplots(figsize=(7.5, 5))
    label_added = False
    for xs, ys in zip(segs_x_PC1, segs_y_PC1):
        ax1.plot(xs, ys, "-", color="darkblue", lw=1.8, zorder=3)
        mi = _marker_idx(xs)
        ax1.plot(xs[mi], ys[mi], "^", color="darkblue", markersize=6,
                 label="PC1" if not label_added else "_nolegend_", zorder=4)
        label_added = True
    ax1.set_xlabel("PC1", fontsize=21, color="black")
    ax1.set_ylabel("Number of Materials", fontsize=21)
    ax1.tick_params(axis="x", colors="black", labelsize=18.5, direction="in", length=8, width=1.5)
    ax1.tick_params(axis="y", labelsize=18.5, direction="in", length=8, width=1.5, right=True, left=True)
    ax1.tick_params(axis="both", which="minor", direction="in", length=4, width=1.5, right=True, left=True)
    ax1.minorticks_on()
    ax1.yaxis.set_minor_locator(AutoMinorLocator(7))
    ax1.xaxis.set_minor_locator(AutoMinorLocator(15))
    all_y = np.concatenate(segs_y_PC1 + segs_y_PC2) if segs_y_PC2 else np.concatenate(segs_y_PC1)
    y_hi  = np.ceil(np.nanmax(all_y) / 50) * 50
    step  = max(50, round(y_hi / 5 / 50) * 50)
    ax1.set_ylim(0, y_hi * 1.08); ax1.set_yticks(np.arange(0, y_hi + 1, step))

    ax2 = ax1.twiny()
    label_added = False
    for xs, ys in zip(segs_x_PC2, segs_y_PC2):
        ax2.plot(xs, ys, "-", color="crimson", lw=1.8, zorder=3)
        mi = _marker_idx(xs)
        ax2.plot(xs[mi], ys[mi], "D", color="crimson", markersize=5,
                 label="PC2" if not label_added else "_nolegend_", zorder=4)
        label_added = True
    ax2.set_xlabel("PC2", fontsize=21, color="black")
    ax2.tick_params(axis="x", colors="black", labelsize=18.5, direction="in", length=8, width=1.5)
    ax2.tick_params(axis="both", colors="black", which="minor", direction="in", length=4, width=1.5)
    ax2.minorticks_on(); ax2.xaxis.set_minor_locator(AutoMinorLocator(15))
    _leg_pc1 = Line2D([0],[0], marker='^', color='darkblue', lw=1.8, markersize=7, linestyle='-', label='PC1: Number of Materials')
    _leg_pc2 = Line2D([0],[0], marker='D', color='crimson',  lw=1.8, markersize=6, linestyle='-', label='PC2: Number of Materials')
    ax1.legend(handles=[_leg_pc1, _leg_pc2], fontsize=14, fancybox=True, edgecolor='k', loc='upper right', framealpha=1, ncol=1)
    plt.tight_layout()
    for fmt in ['png', 'pdf']:
        plt.savefig(f"{save_base}_NumMaterials_vs_PC1_PC2.{fmt}", dpi=600, bbox_inches="tight")
    plt.close()
    log.info(f"Material count plot saved: {save_base}_NumMaterials_vs_PC1_PC2.png/pdf")


def plot_theta_d_statistics_vs_components(df, save_base, symbol_converter, dt_col='DT'):
    log.info(f"Generating {dt_col} statistics vs PC1/PC2 plot ...")
    y_label      = symbol_converter.get_axis_label(dt_col)
    sorted_data  = df.sort_values("PC1").reset_index(drop=True)
    bin_size_PC1 = 0.950; bin_size_PC2 = 0.475
    bins_PC1 = np.arange(sorted_data["PC1"].min(), sorted_data["PC1"].max() + bin_size_PC1, bin_size_PC1)
    bins_PC2 = np.arange(sorted_data["PC2"].min(), sorted_data["PC2"].max() + bin_size_PC2, bin_size_PC2)

    def agg_stats(grp_col, mean_col, bins):
        return sorted_data.groupby(pd.cut(sorted_data[grp_col], bins=bins)).agg(
            **{mean_col: (grp_col, "mean"), "mean_theta": (dt_col, "mean"),
               "std_theta": (dt_col, "std"),
               "q10_theta": (dt_col, lambda x: np.percentile(x, 10) if len(x) > 0 else np.nan),
               "q90_theta": (dt_col, lambda x: np.percentile(x, 90) if len(x) > 0 else np.nan),
               "count": (dt_col, "size")}
        ).reset_index()

    grouped_PC1 = agg_stats("PC1", "mean_PC1", bins_PC1)
    grouped_PC2 = agg_stats("PC2", "mean_PC2", bins_PC2)

    fig, ax1 = plt.subplots(figsize=(7.5, 5))
    ax1.errorbar(grouped_PC1["mean_PC1"], grouped_PC1["mean_theta"],
                 yerr=grouped_PC1["std_theta"], fmt='^-', color="darkblue",
                 ecolor="darkblue", elinewidth=1.5, capsize=6, zorder=10,
                 label=r"PC1: $\Theta_D$ Mean $\pm$ Std")
    ax1.fill_between(grouped_PC1["mean_PC1"], grouped_PC1["q10_theta"],
                     grouped_PC1["q90_theta"], color="blue", alpha=0.2, zorder=5)
    ax1.set_xlabel("PC1", fontsize=21, color="black"); ax1.set_ylabel(y_label, fontsize=21)
    ax1.tick_params(axis="x", colors="black", labelsize=18.5, direction='in', length=8, width=1.5)
    ax1.tick_params(axis="y", labelsize=18.5, direction='in', length=8, right=True, left=True, width=1.5)
    ax1.tick_params(axis='both', which='minor', direction='in', length=4, right=True, left=True, width=1.5)
    ax1.minorticks_on(); ax1.yaxis.set_minor_locator(AutoMinorLocator(7)); ax1.xaxis.set_minor_locator(AutoMinorLocator(15))
    ax2 = ax1.twiny()
    ax2.errorbar(grouped_PC2["mean_PC2"], grouped_PC2["mean_theta"],
                 yerr=grouped_PC2["std_theta"], fmt='D-', color="crimson",
                 ecolor="crimson", elinewidth=1.5, capsize=3, zorder=10,
                 label=r"PC2: $\Theta_D$ Mean $\pm$ Std")
    ax2.fill_between(grouped_PC2["mean_PC2"], grouped_PC2["q10_theta"],
                     grouped_PC2["q90_theta"], color="red", alpha=0.2, zorder=5)
    ax2.set_xlabel("PC2", fontsize=21, color="black")
    ax2.tick_params(axis="x", colors="black", labelsize=18.5, direction='in', length=8, width=1.5)
    ax2.tick_params(axis='both', colors="black", which='minor', direction='in', length=4, width=1.5)
    ax2.minorticks_on(); ax2.xaxis.set_minor_locator(AutoMinorLocator(15))
    h1, l1 = ax1.get_legend_handles_labels(); h2, l2 = ax2.get_legend_handles_labels()
    ax1.set_ylim(0, 3000)
    ax1.legend(h1 + h2, l1 + l2, fontsize=14, fancybox=True, edgecolor="k",
               loc="upper center", framealpha=1, ncol=2)
    plt.tight_layout()
    for fmt in ['png', 'pdf']:
        plt.savefig(f"{save_base}_ThetaD_Stats_vs_PC1_PC2.{fmt}", dpi=600, bbox_inches="tight")
    plt.close()
    log.info(f"Theta_D statistics plot saved: {save_base}_ThetaD_Stats_vs_PC1_PC2.png/pdf")


def plot_theta_d_distribution(df, save_base, symbol_converter, dt_col='DT'):
    log.info(f"Generating {dt_col} distribution plot ...")
    x_label     = symbol_converter.get_axis_label(dt_col)
    sorted_data = df.sort_values(dt_col).reset_index(drop=True)
    bin_size    = 75
    bins        = np.arange(sorted_data[dt_col].min(), sorted_data[dt_col].max() + bin_size, bin_size)
    sorted_data["bin"] = pd.cut(sorted_data[dt_col], bins=bins, labels=bins[:-1])
    grouped_total = sorted_data.groupby("bin").size().reset_index(name="count")
    grouped_total["mean_theta"] = grouped_total["bin"].astype(float)
    grouped_total["count"]      = grouped_total["count"].replace(0, np.nan)

    def gradient_bar(ax, x, height, width=60, color_top="#B199FF", color_bottom="#6F00FF"):
        cmap   = LinearSegmentedColormap.from_list("bar_cmap", [color_bottom, color_top])
        grad   = np.linspace(0, 1, 250)
        grad2d = np.outer(grad, grad)
        ax.imshow(grad2d, extent=(x - width/2, x + width/2, 0, height),
                  aspect="auto", origin="lower", cmap=cmap, zorder=1, alpha=1, interpolation="bicubic")
        ax.plot([x-width/2, x+width/2, x+width/2, x-width/2, x-width/2],
                [0, 0, height, height, 0], color="#4B0082", lw=1.3, zorder=2)

    def smooth_segments(x, y, points=500):
        segs_x, segs_y = [], []
        is_valid = ~np.isnan(y); start = None
        for i, valid in enumerate(is_valid):
            if valid and start is None: start = i
            elif not valid and start is not None:
                if i - start > 1:
                    xs, ys = x[start:i], y[start:i]
                    xnew = np.linspace(xs.min(), xs.max(), points)
                    segs_x.append(xnew); segs_y.append(make_interp_spline(xs, ys, k=3)(xnew))
                start = None
        if start is not None and len(x) - start > 1:
            xs, ys = x[start:], y[start:]
            xnew = np.linspace(xs.min(), xs.max(), points)
            segs_x.append(xnew); segs_y.append(make_interp_spline(xs, ys, k=3)(xnew))
        return segs_x, segs_y

    fig, ax = plt.subplots(figsize=(7, 4.5))
    for _, row in grouped_total.iterrows():
        if not np.isnan(row["count"]):
            gradient_bar(ax, row["mean_theta"], row["count"], width=55)
    x = grouped_total["mean_theta"].values; y = grouped_total["count"].values
    for xs, ys in zip(*smooth_segments(x, y)):
        ax.plot(xs, ys, color="orangered", linewidth=3.5, linestyle='-', zorder=3)
    ax.set_xlabel(x_label, fontsize=18); ax.set_ylabel("Number of Materials", fontsize=18)
    ax.tick_params(axis='both', labelsize=16)
    ax.tick_params(axis='both', which='both', direction='in', length=8, width=1.5,
                   top=True, bottom=True, left=True, right=True)
    ax.tick_params(axis='both', which='minor', direction='in', length=3.5, width=1.5,
                   top=True, bottom=True, left=True, right=True)
    ax.minorticks_on(); ax.xaxis.set_minor_locator(AutoMinorLocator(15)); ax.yaxis.set_minor_locator(AutoMinorLocator(7))
    plt.xlim(-100, 2300); plt.tight_layout()
    for fmt in ['png', 'pdf']:
        plt.savefig(f"{save_base}_Count_vs_ThetaD.{fmt}", dpi=600, bbox_inches="tight")
    plt.close()
    log.info(f"Theta_D distribution plot saved: {save_base}_Count_vs_ThetaD.png/pdf")


def _build_contour_grid(x, y, z, resolution=150):
    xi = np.linspace(x.min(), x.max(), resolution)
    yi = np.linspace(y.min(), y.max(), resolution)
    xi, yi = np.meshgrid(xi, yi)
    zi = griddata((x, y), z, (xi, yi), method='linear')
    return xi, yi, np.clip(zi, np.nanmin(z), np.nanmax(z))

def _draw_single_contour(ax, xi, yi, zi, label, cmap='jet'):
    levels  = np.linspace(np.nanmin(zi), np.nanmax(zi), 1000)
    contour = ax.contourf(xi, yi, zi, levels=levels, cmap=cmap)
    ax.axhline(0, color='black', linewidth=1.5, linestyle='--')
    ax.axvline(0, color='black', linewidth=1.5, linestyle='--')
    ax.set_xlabel('Principal Component 1', fontsize=24)
    ax.set_ylabel('Principal Component 2', fontsize=24)
    ax.tick_params(axis='both', which='both', direction='in', labelsize=24,
                   top=True, bottom=True, left=True, right=True, length=8)
    ax.tick_params(axis='both', which='minor', direction='in', length=3, width=1.0)
    ax.minorticks_on(); ax.xaxis.set_minor_locator(AutoMinorLocator(5)); ax.yaxis.set_minor_locator(AutoMinorLocator(5))
    cbar = plt.colorbar(contour, ax=ax)
    cbar.set_label(label, fontsize=24)
    cbar.set_ticks(np.linspace(np.nanmin(zi), np.nanmax(zi), 6))
    cbar.ax.tick_params(labelsize=24)

def plot_contours(df, feature_cols, save_base, symbol_converter):
    available = [f for f in feature_cols if f in df.columns]
    if not available: log.warning("No feature columns for contour plots."); return
    x = df['PC1'].values; y = df['PC2'].values
    grids  = {feat: _build_contour_grid(x, y, df[feat].values) for feat in available}
    labels = {feat: symbol_converter.get_axis_label(feat) for feat in available}
    n = len(available); ncols = min(3, n); nrows = int(np.ceil(n / ncols))
    fig, axes = plt.subplots(nrows, ncols, figsize=(7 * ncols, 4.5 * nrows))
    axes_flat = np.array(axes).flatten()
    for idx, feat in enumerate(available):
        _draw_single_contour(axes_flat[idx], *grids[feat], labels[feat])
    for idx in range(n, len(axes_flat)):
        axes_flat[idx].set_visible(False)
    plt.suptitle("PCA Space — Property Contour Maps", fontsize=18, fontweight='bold', y=1.01)
    plt.tight_layout()
    for fmt in ['png', 'pdf']:
        plt.savefig(f"{save_base}_Contours_Combined.{fmt}", dpi=600, bbox_inches="tight")
    plt.close()
    log.info(f"Combined contour panel saved: {save_base}_Contours_Combined.png/pdf")


def plot_property_threshold_analysis(df, save_base, symbol_converter, dt_col='DT'):
    log.info("Generating property threshold analysis plot ...")
    def gradient_fill(ax, x, y_min, y_max, cmap_name, vmin, vmax):
        if not len(x): return
        norm = mcolors.Normalize(vmin=vmin, vmax=vmax)
        cmap = cmap_name if callable(cmap_name) else plt.get_cmap(cmap_name)
        for j in range(len(x) - 1):
            x0, x1 = x[j], x[j+1]
            ylo = (y_min[j] + y_min[j+1]) / 2; yhi = (y_max[j] + y_max[j+1]) / 2
            ax.fill_between([x0, x1], [ylo, ylo], [yhi, yhi],
                            color=cmap(norm(x[j])), edgecolor='none')

    x_label        = f"Debye Temperature Threshold ({COLUMN_UNITS.get(dt_col, 'K')})"
    dt_thresholds_low  = np.arange(-1000, 350 + 1, 2)
    dt_thresholds_high = np.arange(1000, df[dt_col].max() + 1, 2)
    input_columns  = [c for c in DataLoader.FEATURE_COLUMNS if c in df.columns and c != dt_col]
    if not input_columns: log.warning("No property columns for threshold analysis."); return
    y_labels = symbol_converter.labels_for_columns(input_columns)

    def collect_ranges(thresholds, mask_fn):
        out = {}
        for thr in thresholds:
            sub = df[mask_fn(thr)]
            if len(sub): out[thr] = sub[input_columns].describe().loc[['min', 'max']]
        return pd.concat(out, axis=0) if out else pd.DataFrame()

    df_low  = collect_ranges(dt_thresholds_low,  lambda t: df[dt_col] < t)
    df_high = collect_ranges(dt_thresholds_high, lambda t: df[dt_col] > t)
    if df_low.empty and df_high.empty: log.warning("No data for threshold analysis."); return
    df_low.index.names  = ['Threshold_K', 'Stat']
    df_high.index.names = ['Threshold_K', 'Stat']
    combined = pd.concat([df_low, df_high])

    fig, axes = plt.subplots(nrows=len(input_columns), ncols=1,
                              figsize=(16.5, 2.6 * len(input_columns)), sharex=True)
    if len(input_columns) == 1: axes = np.array([axes])

    for i, feature in enumerate(input_columns):
        ax = axes[i]
        if feature not in combined.columns: continue
        min_vals   = combined.xs('min', level='Stat')[feature]
        max_vals   = combined.xs('max', level='Stat')[feature]
        thresholds = min_vals.index
        low_mask   = thresholds < 350
        mid_mask   = (thresholds >= 350) & (thresholds <= 1000)
        high_mask  = thresholds > 1000
        cmap_low   = LinearSegmentedColormap.from_list('#80acff', ['#2276FE', '#92CFFF'])
        cmap_high  = LinearSegmentedColormap.from_list('#fc7c82', ['#FF737F', '#FF2F3B'])
        ax.plot(thresholds[low_mask],  min_vals[low_mask],  color='blue',    linestyle='--', linewidth=2)
        ax.plot(thresholds[low_mask],  max_vals[low_mask],  color='darkblue', linestyle='-', linewidth=2)
        ax.fill_between(thresholds[mid_mask], min_vals[mid_mask], max_vals[mid_mask], color='white', edgecolor='none')
        ax.plot(thresholds[mid_mask],  min_vals[mid_mask],  color='black',   linestyle='--', linewidth=2)
        ax.plot(thresholds[mid_mask],  max_vals[mid_mask],  color='black',   linestyle='-',  linewidth=2)
        if np.any(low_mask):
            x_l = thresholds[low_mask].values
            gradient_fill(ax, x_l, min_vals[low_mask].values, max_vals[low_mask].values, cmap_low, x_l.min(), x_l.max())
        if np.any(high_mask):
            x_h = thresholds[high_mask].values
            gradient_fill(ax, x_h, min_vals[high_mask].values, max_vals[high_mask].values, cmap_high, x_h.min(), x_h.max())
        ax.plot(thresholds[high_mask], min_vals[high_mask], color='red',     linestyle='--', linewidth=2)
        ax.plot(thresholds[high_mask], max_vals[high_mask], color='darkred', linestyle='-',  linewidth=2)
        if np.any(low_mask) and np.any(mid_mask):
            le = thresholds[low_mask][-1]
            ax.vlines(x=le, ymin=min_vals[le], ymax=max_vals[le], color='black', linestyle='-', linewidth=2)
        if np.any(high_mask) and np.any(mid_mask):
            hs = thresholds[high_mask][0]
            ax.vlines(x=hs, ymin=min_vals[hs], ymax=max_vals[hs], color='black', linestyle='-', linewidth=2)
        ax.set_ylabel(y_labels[i], fontsize=25)
        ax.tick_params(labelsize=25, axis='both', which='both', direction='in',
                       length=10, top=True, bottom=True, left=True, right=True, width=1.8)
        ax.tick_params(axis='both', which='minor', direction='in', length=5, width=1.8,
                       top=True, bottom=True, left=True, right=True)
        ax.minorticks_on(); ax.xaxis.set_minor_locator(AutoMinorLocator(15)); ax.yaxis.set_minor_locator(AutoMinorLocator(5))
        ax.yaxis.set_major_locator(MaxNLocator(3)); ax.xaxis.set_major_locator(MaxNLocator(8))
        for spine in ax.spines.values(): spine.set_linewidth(2.0)

    max_handle = TwoColorLine('darkblue', 'darkred', linestyle='-',  linewidth=2.5)
    min_handle = TwoColorLine('blue',     'red',     linestyle='--', linewidth=2.5)
    fill_high  = Patch(color='#fc7c82'); fill_low = Patch(color='#80acff')
    dt_sym     = symbol_converter.get(dt_col)
    axes[0].legend(
        handles=[max_handle, min_handle, fill_high, fill_low],
        labels=['Max Value Line', 'Min Value Line',
                f'Higher {dt_sym} Range', f'Lower {dt_sym} Range'],
        handler_map={TwoColorLine: HandlerTwoColorLine(gap_ratio=0.08)},
        loc='upper left', fontsize=15, frameon=True, edgecolor='black',
    )
    axes[-1].set_xlabel(x_label, fontsize=25, labelpad=12)
    plt.tight_layout(rect=[0, 0.03, 1, 0.97])
    for fmt in ['png', 'pdf']:
        plt.savefig(f"{save_base}_Property_Threshold_Analysis.{fmt}", dpi=600, bbox_inches="tight")
    plt.close()
    log.info(f"Property threshold analysis saved: {save_base}_Property_Threshold_Analysis.png/pdf")

    low_mat  = df[df[dt_col] < 350]; high_mat = df[df[dt_col] >= 1000]
    log.info(f"Materials in Low  {dt_col} regime (<350 K):  {len(low_mat):,}")
    log.info(f"Materials in High {dt_col} regime (>=1000 K): {len(high_mat):,}")


# ============================================================================
# SUMMARY PRINTER
# ============================================================================
def print_category_summary(df):
    counts = df["Category"].value_counts()
    total  = len(df)
    log.info("=" * 60)
    log.info("CATEGORY SUMMARY")
    log.info(f"  {'Category':<35}  {'Count':>8}  {'Share':>7}")
    log.info(f"  {'-'*35}  {'-'*8}  {'-'*7}")
    for cat in CategoryConfig.DISPLAY_ORDER:
        if cat not in counts: continue
        count = counts[cat]; pct = count / total * 100
        log.info(f"  {cat:<35}  {count:>8,}  {pct:>6.1f}%")
    log.info(f"  {'-'*55}")
    log.info(f"  Total: {total:,} materials")
    log.info("=" * 60)


# ============================================================================
# EXCEL EXPORT  (identical to source — Colors.* -> log.*)
# ============================================================================
def export_to_excel(df_pca, loadings, runner, feature_cols, symbol_converter,
                    save_base, dt_col='DT', timing_log=None, n_cpus=1, backend_label='N/A'):
    log.info("Preparing Excel export ...")
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    TAB_COLORS = ["1F4E79","2E75B6","2F5496","375623","548235","70AD47",
                  "833C00","C55A11","7B7B7B","595959","203864","BF9000"]
    FILLS = {
        'raw':    ("1F4E79","FFFFFF"), 'pca':    ("2E75B6","FFFFFF"),
        'stats':  ("375623","FFFFFF"), 'dist':   ("833C00","FFFFFF"),
        'thresh': ("7B7B7B","FFFFFF"), 'corr':   ("2F5496","FFFFFF"),
        'contrib':("203864","FFFFFF"),
    }

    def make_ws(title, tab_idx):
        ws = wb.create_sheet(title=title)
        ws.sheet_properties.tabColor = TAB_COLORS[tab_idx % len(TAB_COLORS)]
        ws.row_dimensions[1].height  = 18
        return ws

    pc_cols_present = [c for c in df_pca.columns if c.startswith('PC')]
    raw_cols        = [c for c in df_pca.columns if not c.startswith('PC')]

    ws1 = make_ws("01 Raw Data", 0)
    _write_df_to_sheet(ws1, df_pca[raw_cols].copy(), header_fill=FILLS['raw'][0], header_font=FILLS['raw'][1])
    log.info("Sheet 01 – Raw Data")

    ws2 = make_ws("02 PCA Scores", 1)
    id_cols = [c for c in ['material_id','formula','Category'] if c in df_pca.columns]
    _write_df_to_sheet(ws2, df_pca[id_cols + pc_cols_present].copy(),
                       header_fill=FILLS['pca'][0], header_font=FILLS['pca'][1], num_fmt='0.00000')
    log.info("Sheet 02 – PCA Scores")

    ws3 = make_ws("03 Explained Variance", 2)
    evr = runner.explained_variance; cum = np.cumsum(evr)
    _write_df_to_sheet(ws3, pd.DataFrame({
        'Principal Component': [f'PC{i+1}' for i in range(len(evr))],
        'Explained Variance':  evr, 'Cumulative Variance': cum,
        'Explained Variance (%)': evr*100, 'Cumulative Variance (%)': cum*100,
    }), header_fill=FILLS['pca'][0], header_font=FILLS['pca'][1])
    log.info("Sheet 03 – Explained Variance")

    ws4 = make_ws("04 PC Loadings", 3)
    n_pcs = loadings.shape[1]
    df_load = pd.DataFrame(loadings, index=feature_cols,
                            columns=[f'PC{i+1}' for i in range(n_pcs)]).reset_index()
    df_load.rename(columns={'index':'Feature'}, inplace=True)
    df_load.insert(1, 'Label', [symbol_converter.get(c) for c in feature_cols])
    _write_df_to_sheet(ws4, df_load, header_fill=FILLS['pca'][0], header_font=FILLS['pca'][1], num_fmt='0.00000')
    log.info("Sheet 04 – PC Loadings")

    ws5 = make_ws("05 Category Summary", 4)
    total = len(df_pca); dt_avail = dt_col in df_pca.columns
    rows_cat = []
    for cat in CategoryConfig.DISPLAY_ORDER:
        sub = df_pca[df_pca['Category'] == cat]
        if sub.empty: continue
        row = {'Category': cat, 'Count': len(sub), 'Share (%)': round(len(sub)/total*100, 2)}
        if dt_avail:
            row.update({f'{dt_col} Min (K)': sub[dt_col].min(), f'{dt_col} Max (K)': sub[dt_col].max(),
                        f'{dt_col} Mean (K)': sub[dt_col].mean(), f'{dt_col} Std (K)': sub[dt_col].std()})
        rows_cat.append(row)
    rows_cat.append({'Category':'TOTAL','Count':total,'Share (%)':100.0})
    df_cat = pd.DataFrame(rows_cat)
    _write_df_to_sheet(ws5, df_cat, header_fill=FILLS['stats'][0], header_font=FILLS['stats'][1], num_fmt='0.00')
    for row_idx in range(2, ws5.max_row + 1):
        cat_name = ws5.cell(row=row_idx, column=1).value
        hex_fill = CategoryConfig.XL_HEADER_FILLS.get(cat_name)
        if hex_fill:
            ws5.cell(row=row_idx, column=1).fill = PatternFill("solid", fgColor=hex_fill)
            ws5.cell(row=row_idx, column=1).font = Font(name='Arial', size=10, bold=True, color="000000")
    log.info("Sheet 05 – Category Summary")

    ws6 = make_ws("06 Feature Statistics", 5)
    stat_rows = []
    for cat in CategoryConfig.DISPLAY_ORDER:
        sub = df_pca[df_pca['Category'] == cat]
        if sub.empty: continue
        for feat in feature_cols:
            if feat not in sub.columns: continue
            stat_rows.append({'Category': cat, 'Feature': feat, 'Label': symbol_converter.get(feat),
                              'Unit': COLUMN_UNITS.get(feat,''), 'Min': sub[feat].min(),
                              'Max': sub[feat].max(), 'Mean': sub[feat].mean(),
                              'Median': sub[feat].median(), 'Std': sub[feat].std(), 'Count': sub[feat].count()})
    _write_df_to_sheet(ws6, pd.DataFrame(stat_rows), header_fill=FILLS['stats'][0], header_font=FILLS['stats'][1])
    log.info("Sheet 06 – Feature Statistics")

    for ws_num, (pc_col, bin_size, tab_idx) in enumerate([('PC1',0.950,6),('PC2',0.475,7)], start=7):
        ws_n = make_ws(f"0{ws_num} {pc_col} Bin Stats", tab_idx)
        df_bins = _compute_pc_bin_stats(df_pca, pc_col, dt_col, bin_size)
        df_bins.rename(columns={'count':'N_materials','DT_mean':f'{dt_col}_mean (K)','DT_std':f'{dt_col}_std (K)',
                                'DT_min':f'{dt_col}_min (K)','DT_max':f'{dt_col}_max (K)',
                                'DT_p10':f'{dt_col}_p10 (K)','DT_p90':f'{dt_col}_p90 (K)'}, inplace=True)
        _write_df_to_sheet(ws_n, df_bins, header_fill=FILLS['dist'][0], header_font=FILLS['dist'][1])
        log.info(f"Sheet 0{ws_num} – {pc_col} Bin Stats")

    ws9 = make_ws("09 DT Distribution", 8)
    _write_df_to_sheet(ws9, _compute_dt_distribution(df_pca, dt_col, 75),
                       header_fill=FILLS['dist'][0], header_font=FILLS['dist'][1], num_fmt='0.00')
    log.info("Sheet 09 – DT Distribution")

    ws10 = make_ws("10 Threshold Low (<350K)", 9)
    ws11 = make_ws("11 Threshold High (>1000K)", 10)
    df_thresh_low, df_thresh_high = _compute_threshold_ranges(df_pca, feature_cols, dt_col, 350, 1000)
    _write_df_to_sheet(ws10, df_thresh_low,  header_fill=FILLS['thresh'][0], header_font=FILLS['thresh'][1])
    _write_df_to_sheet(ws11, df_thresh_high, header_fill=FILLS['thresh'][0], header_font=FILLS['thresh'][1])
    log.info("Sheet 10 – Threshold Low  |  Sheet 11 – Threshold High")

    ws12 = make_ws("12 Correlation Matrix", 11)
    corr_cols = [c for c in feature_cols + pc_cols_present if c in df_pca.columns]
    df_corr   = df_pca[corr_cols].corr().reset_index().rename(columns={'index':'Feature'})
    _write_df_to_sheet(ws12, df_corr, header_fill=FILLS['corr'][0], header_font=FILLS['corr'][1])
    for row_idx in range(2, ws12.max_row + 1):
        for col_idx in range(2, ws12.max_column + 1):
            cell = ws12.cell(row=row_idx, column=col_idx); val = cell.value
            if not isinstance(val, (float, int, np.floating)): continue
            t = (float(val) + 1) / 2
            r = int(255*(1-t)); g = int(255*t); b = 120
            cell.fill = PatternFill("solid", fgColor=f"{r:02X}{g:02X}{b:02X}")
            cell.font = Font(name='Arial', size=9, color="000000" if 0.25 < t < 0.75 else "FFFFFF")
    log.info("Sheet 12 – Correlation Matrix (heatmap)")

    ws13 = make_ws("13 PCA Contributions", 0)
    contrib = (loadings**2) * runner.explained_variance[np.newaxis,:] * 100
    df_contrib = pd.DataFrame(contrib, index=feature_cols,
                               columns=[f'PC{i+1} contrib (%)' for i in range(n_pcs)])
    df_contrib['Total contrib (%)'] = df_contrib.sum(axis=1)
    df_contrib.insert(0,'Label',[symbol_converter.get(c) for c in feature_cols])
    df_contrib = df_contrib.reset_index().rename(columns={'index':'Feature'})
    _write_df_to_sheet(ws13, df_contrib, header_fill=FILLS['contrib'][0], header_font=FILLS['contrib'][1], num_fmt='0.000')
    max_c = float(df_contrib[[c for c in df_contrib.columns if 'contrib' in c]].max().max())
    for row_idx in range(2, ws13.max_row+1):
        for col_idx in range(3, ws13.max_column+1):
            cell = ws13.cell(row=row_idx, column=col_idx); val = cell.value
            if not isinstance(val,(float,int,np.floating)) or max_c == 0: continue
            intensity = min(1.0, float(val)/max_c)
            b_val = int(220 - intensity*130); g_val = int(235 - intensity*50)
            cell.fill = PatternFill("solid", fgColor=f"00{g_val:02X}{b_val:02X}")
    log.info("Sheet 13 – PCA Contributions")

    if timing_log:
        ws14 = make_ws("14 Timing Report", 11)
        df_timing = pd.DataFrame([{
            'Phase': k, 'Time (s)': round(v,4),
            'Time (mm:ss)': f"{int(v//60):02d}:{v%60:05.2f}",
            'Share of total (%)': round(v/timing_log.get('TOTAL wall-clock',v)*100, 2),
        } for k,v in timing_log.items()])
        _write_df_to_sheet(ws14, df_timing, header_fill="2F4F4F", header_font="FFFFFF")

        try:
            cpu_p = psutil.cpu_count(logical=False) if psutil else 'N/A'
            cpu_l = psutil.cpu_count(logical=True)  if psutil else 'N/A'
            ram   = round(psutil.virtual_memory().total/1e9, 2) if psutil else 'N/A'
        except Exception:
            cpu_p = cpu_l = ram = 'N/A'
        try:
            import sklearn, scipy, openpyxl as oxl, seaborn as sns_mod
            sklearn_ver=sklearn.__version__; scipy_ver=scipy.__version__
            openpyxl_ver=oxl.__version__; seaborn_ver=sns_mod.__version__
        except Exception:
            sklearn_ver=scipy_ver=openpyxl_ver=seaborn_ver='N/A'
        try:
            import plotly; plotly_ver=plotly.__version__
        except ImportError:
            plotly_ver='not installed'

        sys_info = [
            ('── Run Info ──',''),
            ('Run timestamp', datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('Total materials', len(df_pca)), ('Features used', ', '.join(feature_cols)),
            ('CPUs requested', n_cpus), ('Backend used', backend_label), ('',''),
            ('── System Info ──',''), ('OS', platform.system()+' '+platform.release()),
            ('Hostname', platform.node()), ('CPU model', platform.processor() or 'N/A'),
            ('Physical CPU cores', cpu_p), ('Logical CPU cores', cpu_l), ('Total RAM (GB)', ram), ('',''),
            ('── SLURM Info ──',''),
            ('SLURM Job ID',        os.environ.get('SLURM_JOB_ID','N/A')),
            ('SLURM Node',          os.environ.get('SLURM_NODELIST','N/A')),
            ('SLURM CPUs per task', os.environ.get('SLURM_CPUS_PER_TASK','N/A')),
            ('SLURM Memory (MB)',   os.environ.get('SLURM_MEM_PER_NODE','N/A')),
            ('SLURM Partition',     os.environ.get('SLURM_JOB_PARTITION','N/A')), ('',''),
            ('── Library Versions ──',''),
            ('Python', platform.python_version()), ('NumPy', np.__version__),
            ('Pandas', pd.__version__), ('Matplotlib', matplotlib.__version__),
            ('Scikit-learn', sklearn_ver), ('SciPy', scipy_ver),
            ('Seaborn', seaborn_ver), ('Openpyxl', openpyxl_ver), ('Plotly', plotly_ver),
            ('Psutil', psutil.__version__ if psutil else 'not installed'),
        ]
        info_row = len(df_timing) + 3
        for offset, (label, value) in enumerate(sys_info):
            cl = ws14.cell(row=info_row+offset, column=1, value=str(label))
            cv = ws14.cell(row=info_row+offset, column=2, value=str(value))
            if str(label).startswith('--'):
                cl.font=Font(bold=True,name='Arial',size=10,color='FFFFFF')
                cl.fill=PatternFill("solid",fgColor="2F4F4F")
                cv.fill=PatternFill("solid",fgColor="2F4F4F")
            else:
                cl.font=Font(bold=True,name='Arial',size=10)
                cv.font=Font(name='Arial',size=10)
        max_share = df_timing['Share of total (%)'].max()
        for row_idx in range(2, len(df_timing)+2):
            cell = ws14.cell(row=row_idx, column=4); val = cell.value
            if isinstance(val,(float,int)) and max_share > 0:
                intensity = min(1.0, float(val)/max_share)
                g_val = int(200 - intensity*100)
                cell.fill = PatternFill("solid", fgColor=f"00{g_val:02X}00")
                cell.font = Font(name='Arial',size=10, color="FFFFFF" if intensity > 0.5 else "000000")
        log.info("Sheet 14 – Timing Report")

    xl_path = f"{save_base}_PCA_Export.xlsx"
    wb.save(xl_path)
    log.info(f"Excel workbook saved: {xl_path}")
    for ws in wb.worksheets: log.info(f"  • {ws.title}")
    return xl_path


def export_figure_source_data(df_pca, loadings, runner, feature_cols,
                               symbol_converter, save_base, dt_col='DT'):
    log.info("Preparing Figure Source Data Excel export ...")
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    FILL = ("2F4F4F","FFFFFF")

    def make_ws(title):
        return wb.create_sheet(title=title)

    pc_cols_present = [c for c in df_pca.columns if c.startswith('PC')]
    evr = runner.explained_variance

    ws = make_ws("Fig_ExplainedVariance")
    _write_df_to_sheet(ws, pd.DataFrame({'Principal_Component':[f'PC{i+1}' for i in range(len(evr))],
        'Explained_Variance': evr, 'Cumulative_Variance': np.cumsum(evr)}),
        header_fill=FILL[0], header_font=FILL[1])

    ws = make_ws("Fig_PC_Loadings")
    n_pcs = loadings.shape[1]
    df_load = pd.DataFrame(loadings, index=feature_cols,
                            columns=[f'PC{i+1}_loading' for i in range(n_pcs)])
    df_load.insert(0,'Feature',feature_cols)
    df_load.insert(1,'Label',[symbol_converter.get(c) for c in feature_cols])
    _write_df_to_sheet(ws, df_load.reset_index(drop=True), header_fill=FILL[0], header_font=FILL[1])

    ws = make_ws("Fig_PCA_Scatter")
    id_cols = [c for c in ['material_id','formula','Category'] if c in df_pca.columns]
    _write_df_to_sheet(ws, df_pca[id_cols+pc_cols_present+feature_cols].copy(),
                       header_fill=FILL[0], header_font=FILL[1])

    sorted_data = df_pca.sort_values("PC1").reset_index(drop=True)
    bin_size_PC1, bin_size_PC2 = 0.950, 0.475
    bins_PC1 = np.arange(sorted_data["PC1"].min(), sorted_data["PC1"].max()+bin_size_PC1, bin_size_PC1)
    bins_PC2 = np.arange(sorted_data["PC2"].min(), sorted_data["PC2"].max()+bin_size_PC2, bin_size_PC2)

    def _bin_cnt_src(data, pc_col, bins, dt_col):
        grp = (data.groupby(pd.cut(data[pc_col], bins=bins))
               .agg(mean_pc=(pc_col,'mean'), count=(dt_col,'size'))
               .reset_index(drop=True))
        grp.rename(columns={'mean_pc':pc_col,'count':'N_materials'}, inplace=True)
        return grp

    ws = make_ws("Fig_Count_vs_PC1_PC2")
    df_cnt_pc1 = _bin_cnt_src(sorted_data,'PC1',bins_PC1,dt_col)
    df_cnt_pc2 = _bin_cnt_src(sorted_data,'PC2',bins_PC2,dt_col)
    max_len = max(len(df_cnt_pc1),len(df_cnt_pc2))
    df_cnt_combined = pd.concat([df_cnt_pc1.reindex(range(max_len)).add_suffix('_PC1bin'),
                                  df_cnt_pc2.reindex(range(max_len)).add_suffix('_PC2bin')], axis=1)
    _write_df_to_sheet(ws, df_cnt_combined, header_fill=FILL[0], header_font=FILL[1])

    def _agg_stats_src(data, pc_col, bins, dt_col):
        grp = data.groupby(pd.cut(data[pc_col], bins=bins)).agg(
            mean_pc=(pc_col,'mean'), mean_DT=(dt_col,'mean'), std_DT=(dt_col,'std'),
            q10_DT=(dt_col, lambda x: np.percentile(x,10) if len(x) else np.nan),
            q90_DT=(dt_col, lambda x: np.percentile(x,90) if len(x) else np.nan),
            count=(dt_col,'size'),
        ).reset_index(drop=True)
        grp.rename(columns={'mean_pc':pc_col}, inplace=True)
        return grp

    ws = make_ws("Fig_ThetaD_Stats_vs_PC")
    df_s_pc1 = _agg_stats_src(sorted_data,'PC1',bins_PC1,dt_col)
    df_s_pc2 = _agg_stats_src(sorted_data,'PC2',bins_PC2,dt_col)
    max_len = max(len(df_s_pc1),len(df_s_pc2))
    _write_df_to_sheet(ws, pd.concat([df_s_pc1.reindex(range(max_len)).add_suffix('_PC1'),
                                       df_s_pc2.reindex(range(max_len)).add_suffix('_PC2')], axis=1),
                       header_fill=FILL[0], header_font=FILL[1])

    ws = make_ws("Fig_ThetaD_Distribution")
    _write_df_to_sheet(ws, _compute_dt_distribution(df_pca, dt_col, 75),
                       header_fill=FILL[0], header_font=FILL[1])

    ws_raw = make_ws("Fig_Contours_RawScores")
    id_cols_c = [c for c in ['material_id','formula','Category'] if c in df_pca.columns]
    contour_cols = id_cols_c + ['PC1','PC2'] + [f for f in feature_cols if f in df_pca.columns]
    _write_df_to_sheet(ws_raw, df_pca[contour_cols].copy(), header_fill=FILL[0], header_font=FILL[1])

    x = df_pca['PC1'].values; y = df_pca['PC2'].values
    resolution = 150
    xi_1d = np.linspace(x.min(),x.max(),resolution)
    yi_1d = np.linspace(y.min(),y.max(),resolution)
    xi_grid, yi_grid = np.meshgrid(xi_1d, yi_1d)
    for feat in [f for f in feature_cols if f in df_pca.columns]:
        ws_interp = make_ws(f"Fig_Contour_{feat}_Grid")
        z  = df_pca[feat].values
        zi = griddata((x,y), z, (xi_grid,yi_grid), method='linear')
        zi_clipped = np.clip(zi, np.nanmin(z), np.nanmax(z))
        df_grid = pd.DataFrame(zi_clipped, index=[float(v) for v in np.round(yi_1d,2)],
                               columns=[float(v) for v in np.round(xi_1d,2)])
        df_grid_out = df_grid.reset_index().rename(columns={'index':'PC2_PC1'})
        _write_df_to_sheet(ws_interp, df_grid_out, header_fill=FILL[0], header_font=FILL[1], num_fmt='0.000000')
        log.info(f"  Contour grid saved: {feat} ({resolution}x{resolution})")

    ws_low  = make_ws("Fig_Threshold_Low_Source")
    ws_high = make_ws("Fig_Threshold_High_Source")
    input_columns = [c for c in DataLoader.FEATURE_COLUMNS if c in df_pca.columns and c != dt_col]
    df_thresh_low, df_thresh_high = _compute_threshold_ranges(df_pca, input_columns, dt_col, 350, 1000)
    _write_df_to_sheet(ws_low,  df_thresh_low,  header_fill=FILL[0], header_font=FILL[1])
    _write_df_to_sheet(ws_high, df_thresh_high, header_fill=FILL[0], header_font=FILL[1])

    src_path = f"{save_base}_Figure_Source_Data.xlsx"
    wb.save(src_path)
    log.info(f"Figure source data saved: {src_path}")
    for ws in wb.worksheets: log.info(f"  • {ws.title}")
    return src_path

# ============================================================================
# MAIN
# ============================================================================
def main():
    parser = argparse.ArgumentParser(description="PCA Analysis of Materials Dataset")
    parser.add_argument('--no-log', action='store_true', help='Skip writing a .log file')
    args, _ = parser.parse_known_args()

    global log

    # ── Timing ────────────────────────────────────────────────────────────────
    t_total_start = time.perf_counter()
    timing = {}

    # ── Config ────────────────────────────────────────────────────────────────
    cfg = load_config()

    save_dir  = cfg['save_dir'] or os.getcwd()
    save_name = cfg['save_name'] or 'pca_results'
    save_dir  = os.path.join(save_dir, "PCA")
    os.makedirs(save_dir, exist_ok=True)
    save_base = os.path.join(save_dir, save_name)

    log_path = None if args.no_log else f"{save_base}.log"
    log = setup_logging(log_path)

    log.info("=" * 60)
    log.info("PCA ANALYSIS OF MATERIALS DATASET — HPC VERSION")
    log.info("=" * 60)
    log.info(f"Config loaded | save_base: {save_base}")

    # ── Environment & parallel backend ────────────────────────────────────────
    env         = _detect_environment()
    backend_cfg = _select_backend(cfg['n_cpus'], env)
    log.info(f"Parallel backend: {backend_cfg['label']}")

    # ── Symbol converter ──────────────────────────────────────────────────────
    sym = SymbolConverter(cfg.get('symbol_file'))
    sym_html = SymbolConverter(cfg.get('symbol_file_html')) if cfg.get('symbol_file_html') else sym

    # ── Load data ─────────────────────────────────────────────────────────────
    t0 = time.perf_counter()
    df = DataLoader.load_folder(
        folder      = cfg['data_folder'],
        file_subset = cfg['file_subset'],
        feature_cols= cfg['feature_columns'] or None,
        n_cpus      = cfg['n_cpus'],
        backend_cfg = backend_cfg,
    )
    timing['Data loading'] = time.perf_counter() - t0

    # ── Resolve feature columns ───────────────────────────────────────────────
    feature_cols = DataLoader.resolve_feature_columns(df, cfg['feature_columns'])
    log.info(f"Feature columns: {feature_cols}")
    sym.report_coverage(feature_cols)

    # ── Category summary ──────────────────────────────────────────────────────
    print_category_summary(df)

    # ── PCA ───────────────────────────────────────────────────────────────────
    t0 = time.perf_counter()
    runner   = PCARunner(n_components=min(5, len(feature_cols)))
    df_pca, loadings = runner.fit_transform(df, feature_cols)
    timing['PCA'] = time.perf_counter() - t0

    feature_labels = sym.labels_for_columns(feature_cols)

    # ── Plots ─────────────────────────────────────────────────────────────────
    t0 = time.perf_counter()

    plot_explained_variance(runner.explained_variance, save_base)
    plot_loadings(loadings, feature_labels, save_base)
    plot_scatter(df_pca, save_base, show_labels=cfg['show_labels'])
    plot_material_count_vs_components(df_pca, save_base)
    plot_theta_d_statistics_vs_components(df_pca, save_base, sym)
    plot_theta_d_distribution(df_pca, save_base, sym)
    plot_contours(df_pca, feature_cols, save_base, sym)
    plot_property_threshold_analysis(df_pca, save_base, sym)

    if cfg['interactive_scatter']:
        plot_scatter_interactive(
            df_pca, feature_cols, sym_html, save_base,
            runner      = runner,
            inject_path = cfg.get('inject_file'),
        )

    timing['Plotting'] = time.perf_counter() - t0

    # ── Excel exports ─────────────────────────────────────────────────────────
    t0 = time.perf_counter()
    timing['TOTAL wall-clock'] = time.perf_counter() - t_total_start   # approx so far
    export_to_excel(
        df_pca, loadings, runner, feature_cols, sym, save_base,
        timing_log    = timing,
        n_cpus        = cfg['n_cpus'],
        backend_label = backend_cfg['label'],
    )
    export_figure_source_data(df_pca, loadings, runner, feature_cols, sym, save_base)
    timing['Excel export'] = time.perf_counter() - t0

    # ── Final timing ──────────────────────────────────────────────────────────
    timing['TOTAL wall-clock'] = time.perf_counter() - t_total_start
    log.info("=" * 60)
    log.info("TIMING SUMMARY")
    for phase, secs in timing.items():
        log.info(f"  {phase:<30} {secs:>8.2f}s")
    log.info("=" * 60)
    log.info("PCA run completed.")


if __name__ == "__main__":
    main()